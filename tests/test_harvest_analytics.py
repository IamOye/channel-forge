"""
Tests for scripts/harvest_analytics.py

All Google API calls are mocked — no real network or credential access.
openpyxl is required for the Excel tests (installed via requirements.txt).
"""

import csv
import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_video(
    video_id: str = "abc123",
    title: str = "Test Video",
    duration_seconds: int = 30,
    views: int = 500,
    likes: int = 10,
    published_at: str = "2026-01-01T00:00:00Z",
) -> "VideoRow":
    from scripts.harvest_analytics import VideoRow, AnalyticsHarvester

    return VideoRow(
        video_id=video_id,
        title=title,
        published_at=published_at,
        duration_seconds=duration_seconds,
        type=AnalyticsHarvester.detect_video_type(duration_seconds),
        views=views,
        likes=likes,
    )


# ---------------------------------------------------------------------------
# parse_iso8601_duration
# ---------------------------------------------------------------------------

class TestParseIso8601Duration:
    def _parse(self, s: str) -> int:
        from scripts.harvest_analytics import AnalyticsHarvester
        return AnalyticsHarvester.parse_iso8601_duration(s)

    def test_seconds_only(self) -> None:
        assert self._parse("PT15S") == 15

    def test_minutes_only(self) -> None:
        assert self._parse("PT1M") == 60

    def test_minutes_and_seconds(self) -> None:
        assert self._parse("PT1M30S") == 90

    def test_hours_only(self) -> None:
        assert self._parse("PT1H") == 3600

    def test_hours_minutes_seconds(self) -> None:
        assert self._parse("PT2H3M4S") == 7384

    def test_days(self) -> None:
        assert self._parse("P1D") == 86_400

    def test_empty_string(self) -> None:
        assert self._parse("") == 0

    def test_zero_duration(self) -> None:
        assert self._parse("P0D") == 0

    def test_invalid_string(self) -> None:
        assert self._parse("not-a-duration") == 0


# ---------------------------------------------------------------------------
# detect_video_type
# ---------------------------------------------------------------------------

class TestDetectVideoType:
    def _detect(self, secs: int) -> str:
        from scripts.harvest_analytics import AnalyticsHarvester
        return AnalyticsHarvester.detect_video_type(secs)

    def test_zero_seconds_is_short(self) -> None:
        assert self._detect(0) == "Short"

    def test_exactly_60_is_short(self) -> None:
        assert self._detect(60) == "Short"

    def test_61_is_long_form(self) -> None:
        assert self._detect(61) == "Long-form"

    def test_15_seconds_short(self) -> None:
        assert self._detect(15) == "Short"

    def test_ten_minutes_long(self) -> None:
        assert self._detect(600) == "Long-form"


# ---------------------------------------------------------------------------
# VideoRow
# ---------------------------------------------------------------------------

class TestVideoRow:
    def test_like_rate_normal(self) -> None:
        from scripts.harvest_analytics import VideoRow
        v = VideoRow(video_id="x", views=100, likes=5)
        assert v.like_rate == pytest.approx(0.05)

    def test_like_rate_zero_views(self) -> None:
        from scripts.harvest_analytics import VideoRow
        v = VideoRow(video_id="x", views=0, likes=10)
        assert v.like_rate == 0.0

    def test_to_dict_has_all_columns(self) -> None:
        from scripts.harvest_analytics import VideoRow, VIDEO_CSV_COLUMNS
        v = VideoRow(video_id="x")
        d = v.to_dict()
        for col in VIDEO_CSV_COLUMNS:
            assert col in d, f"Missing column: {col}"

    def test_type_populated_by_detect(self) -> None:
        v = _make_video(duration_seconds=30)
        assert v.type == "Short"

    def test_type_long_form(self) -> None:
        v = _make_video(duration_seconds=120)
        assert v.type == "Long-form"


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

class TestSaveChannelCSV:
    def test_creates_file_with_header(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester, ChannelStats, CHANNEL_CSV_COLUMNS

        h = AnalyticsHarvester(output_dir=tmp_path)
        stats = ChannelStats(date="2026-01-01", subscribers=100, total_views=5000)
        h._save_channel_csv(stats, "2026-01-01")

        csv_path = tmp_path / "channel_stats_2026-01-01.csv"
        assert csv_path.exists()

        with open(csv_path, newline="") as f:
            reader = csv.DictReader(f)
            assert list(reader.fieldnames) == CHANNEL_CSV_COLUMNS

    def test_row_values_correct(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester, ChannelStats

        h = AnalyticsHarvester(output_dir=tmp_path)
        stats = ChannelStats(date="2026-01-01", subscribers=200, total_views=9999, total_videos=8)
        h._save_channel_csv(stats, "2026-01-01")

        with open(tmp_path / "channel_stats_2026-01-01.csv", newline="") as f:
            rows = list(csv.DictReader(f))

        assert len(rows) == 1
        assert rows[0]["subscribers"] == "200"
        assert rows[0]["total_views"] == "9999"

    def test_appends_on_second_call(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester, ChannelStats

        h = AnalyticsHarvester(output_dir=tmp_path)
        h._save_channel_csv(ChannelStats(date="2026-01-01"), "2026-01-01")
        h._save_channel_csv(ChannelStats(date="2026-01-01"), "2026-01-01")

        with open(tmp_path / "channel_stats_2026-01-01.csv", newline="") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2


class TestSaveVideoCSV:
    def test_creates_file_with_correct_columns(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester, VIDEO_CSV_COLUMNS

        h = AnalyticsHarvester(output_dir=tmp_path)
        videos = [_make_video("v1"), _make_video("v2")]
        h._save_video_csv(videos, "2026-01-01")

        csv_path = tmp_path / "video_metrics_2026-01-01.csv"
        assert csv_path.exists()

        with open(csv_path, newline="") as f:
            reader = csv.DictReader(f)
            assert list(reader.fieldnames) == VIDEO_CSV_COLUMNS

    def test_one_row_per_video(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester(output_dir=tmp_path)
        videos = [_make_video(f"vid{i}") for i in range(5)]
        h._save_video_csv(videos, "2026-01-01")

        with open(tmp_path / "video_metrics_2026-01-01.csv", newline="") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 5

    def test_empty_video_list(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester(output_dir=tmp_path)
        h._save_video_csv([], "2026-01-01")

        csv_path = tmp_path / "video_metrics_2026-01-01.csv"
        assert csv_path.exists()

        with open(csv_path, newline="") as f:
            rows = list(csv.DictReader(f))
        assert rows == []

    def test_shorts_type_written(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester(output_dir=tmp_path)
        h._save_video_csv([_make_video("v1", duration_seconds=30)], "2026-01-01")

        with open(tmp_path / "video_metrics_2026-01-01.csv", newline="") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["type"] == "Short"

    def test_long_form_type_written(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester(output_dir=tmp_path)
        h._save_video_csv([_make_video("v1", duration_seconds=120)], "2026-01-01")

        with open(tmp_path / "video_metrics_2026-01-01.csv", newline="") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["type"] == "Long-form"

    def test_output_dir_created_if_missing(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        out = tmp_path / "deep" / "subdir"
        h = AnalyticsHarvester(output_dir=out)
        out.mkdir(parents=True, exist_ok=True)
        h._save_video_csv([], "2026-01-01")
        assert out.exists()


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

class TestSaveExcel:
    def _make_harvester(self, tmp_path: Path):
        from scripts.harvest_analytics import AnalyticsHarvester
        return AnalyticsHarvester(channel_key="money_debate", output_dir=tmp_path)

    def _make_stats(self):
        from scripts.harvest_analytics import ChannelStats
        return ChannelStats(date="2026-01-01", subscribers=14, total_views=3847, total_videos=8)

    def test_excel_has_three_sheets(self, tmp_path: Path) -> None:
        import openpyxl

        h = self._make_harvester(tmp_path)
        videos = [_make_video(f"v{i}", views=i * 100) for i in range(3)]
        xl_path = h._save_excel(self._make_stats(), videos, "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        assert len(wb.sheetnames) == 3

    def test_sheet_names_correct(self, tmp_path: Path) -> None:
        import openpyxl

        h = self._make_harvester(tmp_path)
        xl_path = h._save_excel(self._make_stats(), [], "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        assert wb.sheetnames == ["Channel Overview", "Video Metrics", "Top Performers"]

    def test_channel_overview_has_subscriber_row(self, tmp_path: Path) -> None:
        import openpyxl

        h = self._make_harvester(tmp_path)
        xl_path = h._save_excel(self._make_stats(), [], "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        ws = wb["Channel Overview"]
        values = [ws.cell(row=r, column=1).value for r in range(1, 8)]
        assert "Subscribers" in values

    def test_video_metrics_sheet_has_header(self, tmp_path: Path) -> None:
        import openpyxl
        from scripts.harvest_analytics import VIDEO_CSV_COLUMNS

        h = self._make_harvester(tmp_path)
        videos = [_make_video("v1")]
        xl_path = h._save_excel(self._make_stats(), videos, "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        ws = wb["Video Metrics"]
        headers = [ws.cell(row=1, column=col).value for col in range(1, len(VIDEO_CSV_COLUMNS) + 1)]
        assert headers == VIDEO_CSV_COLUMNS

    def test_video_metrics_sorted_by_views_desc(self, tmp_path: Path) -> None:
        import openpyxl
        from scripts.harvest_analytics import VIDEO_CSV_COLUMNS

        h = self._make_harvester(tmp_path)
        videos = [
            _make_video("low",  views=100),
            _make_video("high", views=5000),
            _make_video("mid",  views=800),
        ]
        xl_path = h._save_excel(self._make_stats(), videos, "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        ws = wb["Video Metrics"]
        views_col = VIDEO_CSV_COLUMNS.index("views") + 1
        row2_views = ws.cell(row=2, column=views_col).value
        assert row2_views == 5000

    def test_auto_filter_enabled(self, tmp_path: Path) -> None:
        import openpyxl

        h = self._make_harvester(tmp_path)
        xl_path = h._save_excel(self._make_stats(), [_make_video("v1")], "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        ws = wb["Video Metrics"]
        assert ws.auto_filter.ref is not None

    def test_top_performers_has_content(self, tmp_path: Path) -> None:
        import openpyxl

        h = self._make_harvester(tmp_path)
        videos = [_make_video(f"v{i}", views=i * 50) for i in range(5)]
        xl_path = h._save_excel(self._make_stats(), videos, "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        ws = wb["Top Performers"]
        # First cell should be "Top 10 by Views"
        assert ws.cell(row=1, column=1).value == "Top 10 by Views"

    def test_conditional_formatting_applied(self, tmp_path: Path) -> None:
        import openpyxl

        h = self._make_harvester(tmp_path)
        videos = [_make_video("v1", views=2000)]
        xl_path = h._save_excel(self._make_stats(), videos, "2026-01-01")

        wb = openpyxl.load_workbook(xl_path)
        ws = wb["Video Metrics"]
        assert len(ws.conditional_formatting._cf_rules) > 0


# ---------------------------------------------------------------------------
# get_channel_name
# ---------------------------------------------------------------------------

class TestGetChannelName:
    def test_money_debate_returns_money_heresy(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester
        h = AnalyticsHarvester(channel_key="money_debate")
        name = h._get_channel_name()
        assert name == "MoneyHeresy"

    def test_unknown_key_falls_back_to_formatted_key(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester
        h = AnalyticsHarvester(channel_key="my_channel")
        # Falls back when channel_key not in CHANNELS
        name = h._get_channel_name()
        assert isinstance(name, str)
        assert len(name) > 0


# ---------------------------------------------------------------------------
# Credentials path
# ---------------------------------------------------------------------------

class TestCredentialsPath:
    def test_path_uses_channel_key(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester(channel_key="money_debate", credentials_dir=".credentials")
        expected = Path(".credentials") / "money_debate_token.json"
        assert h._credentials_path() == expected

    def test_path_uses_custom_credentials_dir(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester(channel_key="test_ch", credentials_dir=tmp_path)
        assert h._credentials_path() == tmp_path / "test_ch_token.json"

    def test_load_credentials_raises_when_file_missing(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester(channel_key="ghost", credentials_dir=tmp_path)
        with pytest.raises(FileNotFoundError):
            h._load_credentials()

    def test_load_credentials_uses_fallback(self, tmp_path: Path) -> None:
        from google.oauth2.credentials import Credentials
        from scripts.harvest_analytics import AnalyticsHarvester

        # Write only the default_token.json
        token_data = {
            "token": "t",
            "refresh_token": "r",
            "token_uri": "https://oauth2.googleapis.com/token",
            "client_id": "cid",
            "client_secret": "sec",
            "scopes": ["https://www.googleapis.com/auth/youtube"],
        }
        (tmp_path / "default_token.json").write_text(json.dumps(token_data))

        h = AnalyticsHarvester(channel_key="any_channel", credentials_dir=tmp_path)
        creds = h._load_credentials()
        assert isinstance(creds, Credentials)

    def test_load_credentials_uses_channel_specific_file(self, tmp_path: Path) -> None:
        from google.oauth2.credentials import Credentials
        from scripts.harvest_analytics import AnalyticsHarvester

        token_data = {
            "token": "tok",
            "refresh_token": "ref",
            "token_uri": "https://oauth2.googleapis.com/token",
            "client_id": "cid",
            "client_secret": "sec",
            "scopes": ["https://www.googleapis.com/auth/youtube"],
        }
        (tmp_path / "money_debate_token.json").write_text(json.dumps(token_data))

        h = AnalyticsHarvester(channel_key="money_debate", credentials_dir=tmp_path)
        creds = h._load_credentials()
        assert isinstance(creds, Credentials)


# ---------------------------------------------------------------------------
# HarvestResult
# ---------------------------------------------------------------------------

class TestHarvestResult:
    def test_summary_format(self) -> None:
        from scripts.harvest_analytics import HarvestResult

        r = HarvestResult(
            channel_key="money_debate",
            videos_count=8,
            total_views=3847,
            subscribers=14,
        )
        s = r.summary()
        assert "8 videos" in s
        assert "3,847" in s
        assert "14 subscribers" in s

    def test_to_dict_has_required_keys(self) -> None:
        from scripts.harvest_analytics import HarvestResult

        r = HarvestResult(channel_key="money_debate")
        d = r.to_dict()
        for key in ("channel_key", "videos_count", "total_views", "subscribers",
                    "output_dir", "saved_files", "is_valid", "errors"):
            assert key in d

    def test_is_valid_false_when_errors(self) -> None:
        from scripts.harvest_analytics import HarvestResult

        r = HarvestResult(channel_key="x", errors=["something broke"], is_valid=False)
        assert r.is_valid is False


# ---------------------------------------------------------------------------
# harvest() top-level function
# ---------------------------------------------------------------------------

class TestHarvestFunction:
    def test_returns_dict(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import harvest

        with patch("scripts.harvest_analytics.AnalyticsHarvester") as MockClass:
            mock_result = MagicMock()
            mock_result.summary.return_value = "Harvested: 0 videos"
            mock_result.to_dict.return_value = {"is_valid": True, "channel_key": "x"}
            MockClass.return_value.harvest.return_value = mock_result

            result = harvest(channel="x", output_dir=tmp_path)

        assert isinstance(result, dict)
        assert result["is_valid"] is True

    def test_returns_error_dict_on_unexpected_exception(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import harvest

        with patch(
            "scripts.harvest_analytics.AnalyticsHarvester",
            side_effect=RuntimeError("boom"),
        ):
            result = harvest(channel="x", output_dir=tmp_path)

        assert result["is_valid"] is False
        assert "error" in result
        assert result["channel_key"] == "x"

    def test_does_not_raise(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import harvest

        with patch(
            "scripts.harvest_analytics.AnalyticsHarvester",
            side_effect=Exception("network down"),
        ):
            # Must not raise
            harvest(channel="x", output_dir=tmp_path)

    def test_passes_channel_key(self, tmp_path: Path) -> None:
        from scripts.harvest_analytics import harvest

        with patch("scripts.harvest_analytics.AnalyticsHarvester") as MockClass:
            mock_result = MagicMock()
            mock_result.summary.return_value = ""
            mock_result.to_dict.return_value = {"is_valid": True}
            MockClass.return_value.harvest.return_value = mock_result

            harvest(channel="money_debate", output_dir=tmp_path)

        MockClass.assert_called_once_with(
            channel_key="money_debate",
            output_dir=tmp_path,
        )


# ---------------------------------------------------------------------------
# Scheduler integration — harvest called from run_daily_analytics
# ---------------------------------------------------------------------------

class TestSchedulerHarvestIntegration:
    def test_harvest_called_from_run_daily_analytics(self) -> None:
        from src.scheduler import run_daily_analytics

        mock_tracker = MagicMock()
        mock_tracker.track_all.return_value = []

        harvest_calls: list[str] = []

        def fake_harvest(channel: str = "money_debate", **kwargs) -> dict:
            harvest_calls.append(channel)
            return {"is_valid": True}

        with patch.dict("sys.modules", {
            "src.analytics.analytics_tracker": MagicMock(AnalyticsTracker=MagicMock(return_value=mock_tracker)),
            "config.channels": MagicMock(CHANNELS=[MagicMock(channel_key="money_debate")]),
            "scripts.harvest_analytics": MagicMock(harvest=fake_harvest),
        }):
            run_daily_analytics()

        assert len(harvest_calls) == 1
        assert harvest_calls[0] == "money_debate"

    def test_harvest_failure_does_not_stop_analytics(self) -> None:
        from src.scheduler import run_daily_analytics

        mock_tracker = MagicMock()
        mock_tracker.track_all.return_value = []

        with patch.dict("sys.modules", {
            "src.analytics.analytics_tracker": MagicMock(AnalyticsTracker=MagicMock(return_value=mock_tracker)),
            "config.channels": MagicMock(CHANNELS=[MagicMock(channel_key="money_debate")]),
            "scripts.harvest_analytics": MagicMock(harvest=MagicMock(side_effect=Exception("api down"))),
        }):
            # Should not raise
            run_daily_analytics()


# ---------------------------------------------------------------------------
# parse_video_item
# ---------------------------------------------------------------------------

class TestParseVideoItem:
    def _make_item(
        self,
        video_id: str = "abc",
        duration: str = "PT30S",
        views: str = "1000",
        likes: str = "50",
        tags: list | None = None,
        description: str = "hello world",
    ) -> dict:
        return {
            "id": video_id,
            "snippet": {
                "title": "My Video",
                "publishedAt": "2026-01-01T00:00:00Z",
                "tags": tags or [],
                "description": description,
                "categoryId": "22",
            },
            "statistics": {
                "viewCount": views,
                "likeCount": likes,
                "commentCount": "5",
                "favoriteCount": "0",
            },
            "contentDetails": {"duration": duration},
        }

    def test_duration_parsed(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester()
        row = h._parse_video_item(self._make_item(duration="PT1M30S"))
        assert row.duration_seconds == 90

    def test_short_type_assigned(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester()
        row = h._parse_video_item(self._make_item(duration="PT30S"))
        assert row.type == "Short"

    def test_long_form_type_assigned(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester()
        row = h._parse_video_item(self._make_item(duration="PT5M"))
        assert row.type == "Long-form"

    def test_tags_joined(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester()
        row = h._parse_video_item(self._make_item(tags=["finance", "money", "shorts"]))
        assert row.tags == "finance,money,shorts"

    def test_description_length_computed(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester()
        row = h._parse_video_item(self._make_item(description="hello world"))
        assert row.description_length == len("hello world")

    def test_views_and_likes_parsed(self) -> None:
        from scripts.harvest_analytics import AnalyticsHarvester

        h = AnalyticsHarvester()
        row = h._parse_video_item(self._make_item(views="2500", likes="75"))
        assert row.views == 2500
        assert row.likes == 75
